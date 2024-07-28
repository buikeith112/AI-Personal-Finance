import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
#from openpyxl.styles import number_format
#from openpyxl.styles import number_format

#load template
wb = openpyxl.load_workbook('Mock.xlsx')

#select sheets
income_sheet = wb['Income']
expense_sheet = wb['Expenses']

#data to add
#example    #date, source, amount
income_data = [
  ('2024-01-01', 'Salary', 2000),
  ('2024-01-15',  'Freelance', 500)
]

# date, category, description, amount
expense_data = [
  ('2024-01-02', 'Groceries', 'Supermarket', 150),
  ('2024-01-05', 'Entertainment', 'Movie Tickets', 30)
]

# function to add data to a sheet
def add_data(sheet, data):
  for row in data:
    sheet.append(row)

# add data to sheets
add_data(income_sheet, income_data)
add_data(expense_sheet, expense_data)

# Format date and currency columns
def format_columns(sheet, date_col_idx, currency_col_idx):
  # date col
  for cell in sheet[date_col_idx]:
    cell.number_format = 'YYYY-MM-DD'
  
  # currency col
  for cell in sheet[get_column_letter(currency_col_idx)]:
    cell.number_format = '$#,##0.00'

# format columns
format_columns(income_sheet, 1, 3)
format_columns(expense_sheet, 1, 4)

wb.save('Trial 1 update.xlsx')
